"""
File loader router supporting PDF, DOCX, XLSX, TXT, and MD files.

Returns a list of LangChain Document objects.
Uses openpyxl for Excel files to prevent external pandas dependencies.
"""

import os
import openpyxl
from langchain_core.documents import Document
from langchain_community.document_loaders import (
    PyPDFLoader,
    Docx2txtLoader,
    TextLoader,
)
from app.utils.logging import get_logger

logger = get_logger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".png", ".jpg", ".jpeg"}


def load_excel_file(file_path: str) -> list[Document]:
    """Parse Excel workbook (.xlsx) into clean text documents using openpyxl."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        docs = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            lines = [f"Sheet Name: {sheet_name}"]
            for row in rows:
                cell_strs = [str(val).strip() for val in row if val is not None and str(val).strip() != ""]
                if cell_strs:
                    lines.append(" | ".join(cell_strs))

            sheet_text = "\n".join(lines)
            if sheet_text.strip():
                docs.append(Document(
                    page_content=sheet_text,
                    metadata={"source": file_path, "sheet_name": sheet_name}
                ))

        logger.info(f"Loaded {len(docs)} sheet(s) from Excel file: {os.path.basename(file_path)}")
        return docs
    except Exception as e:
        logger.error(f"Failed to parse Excel file {file_path} with openpyxl: {e}")
        return []


def load_file(file_path: str) -> list[Document]:
    """Route a file to the appropriate document loader. Returns list of Documents."""
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".pdf":
            return PyPDFLoader(file_path).load()
        elif ext == ".docx":
            return Docx2txtLoader(file_path).load()
        elif ext == ".xlsx":
            return load_excel_file(file_path)
        elif ext in (".txt", ".md"):
            return TextLoader(file_path, encoding="utf-8").load()
        elif ext in (".png", ".jpg", ".jpeg"):
            return []  # Raw standalone images are handled by image_extraction pipeline directly
        else:
            logger.warning(f"Unsupported file extension: {ext}")
            return []
    except Exception as e:
        logger.error(f"Failed to load file {file_path}: {e}")
        return []
